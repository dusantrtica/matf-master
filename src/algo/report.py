from collections import defaultdict
from functional import seq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from src.algo.cp_solver import SimpleCPSolver
from src.algo.data import (
    Session, Group, print_group, split_students_into_groups, print_session, GROUP_SIZE,
)
from src.algo.model import SchedulingInput


def export_schedule_to_excel(solver: SimpleCPSolver,
                             scheduling_input: SchedulingInput,
                             output_path: str):
    """
    Exportuje rešenje u Excel fajl sa jednim tabom po grupi.

    Svaka tab je tablica rasporeda:
      - Redovi    = radni dani  (Ponedeljak, Utorak, ...)
      - Kolone    = sati       (8, 9, 10, ...)
      - Celije    = informacije o sesiji  (naziv predmeta, tip, ucionica)
    """
    variables = solver.get_solution_variables()
    sessions = solver.sessions
    day_names = scheduling_input.settings.working_days
    hours = solver.working_hours
    classrooms = scheduling_input.classrooms
    groups = split_students_into_groups(
        scheduling_input.students_enrolled, GROUP_SIZE
    )

    # Napravimo lookup: group_id -> lista (session, assignment)
    group_entries = defaultdict(list)
    for session, var in zip(sessions, variables):
        group_entries[session.group_id].append((session, var))

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                              fill_type="solid")
    cell_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for group_id in sorted(group_entries.keys()):
        entries = group_entries[group_id]

        sheet_name = print_group(seq(groups).find(lambda grp: grp.id == group_id), scheduling_input.departments)
        print("SHeet Name = ", sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        # Header row: hour slots
        ws.cell(row=1, column=1, value="Day / Hour").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border
        for col_idx, hour in enumerate(hours):
            cell = ws.cell(row=1, column=col_idx + 2, value=f"{hour}:00")
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Kolona sa nazivima dana + celije
        for row_idx, day_name in enumerate(day_names):
            cell = ws.cell(row=row_idx + 2, column=1, value=day_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

            for col_idx in range(len(hours)):
                ws.cell(row=row_idx + 2, column=col_idx + 2).border = thin_border

        # Popuni sesije
        # Napravimo tablicu: (day_index, hour_index) -> lista celija
        grid = defaultdict(list)
        for session, var in entries:
            room = classrooms[var["room"]]
            label = print_session(
                session, groups,
                scheduling_input.courses,
                scheduling_input.departments,
                room_name=room.name,
            )
            grid[(var["day"], var["hour"])].append(label)

        for (day_idx, hour_idx), labels in grid.items():
            cell = ws.cell(
                row=day_idx + 2,
                column=hour_idx + 2,
                value="\n---\n".join(labels),
            )
            cell.alignment = cell_align
            cell.border = thin_border

        # Automatski podesimo širinu kolona
        ws.column_dimensions["A"].width = 16
        for col_idx in range(len(hours)):
            col_letter = chr(ord("B") + col_idx)
            ws.column_dimensions[col_letter].width = 28

        # Podesimo visinu redova
        for row_idx in range(2, 2 + len(day_names)):
            ws.row_dimensions[row_idx].height = 60

    wb.save(output_path)
    return output_path
